from collections import OrderedDict

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, column_index_from_string


# Custom Errors
class NonePointer(Exception):
    """
    Raise when something points on None
    """
    pass


class UndefinedHeaderError(Exception):
    """
    Raise when referenced to non-existing header of exel file
    """
    pass


class InputError(Exception):
    """
    Raise when user input causes error
    """
    pass


# Static Service functions
def get_number(st: str) -> tuple:
    """
    Checks if given parameter is number, returns true if it is (number can be divided by comma or by point) returns True
    if number and number in python format as second argument if it is not number returns unchanged parameter,
    also returns is_changed status if parameter changed
    :param st: str
    :return: tuple(is_num, is_changed, num)
    """
    try:
        tmp = float(st)
        return True, False, tmp
    except (ValueError, TypeError):
        try:
            list_num = st.split(',')
            if len(list_num) == 2:
                neg = False
                if list_num[0].startswith('-'):
                    neg = True
                    list_num[0] = list_num[0][1:]
                if list_num[0].isdigit() and list_num[1].isdigit():
                    if neg:
                        return True, True, -1 * (float(list_num[0] + '.' + list_num[1]))
                    else:
                        return True, True, float(list_num[0] + '.' + list_num[1])
            else:
                return False, False, st
        except AttributeError:
            return False, False, st


def trunc(num: float, precision: int) -> float:
    """
    Returns value truncated to precision signs after comma
    :param precision: int
    :param num: float
    :return: float
    """
    formation = '{:.' + str(precision) + 'f}'
    return float(formation.format(num))


def is_integer(num) -> bool:
    """
    Checks if given number is integer
    :param num: int, float
    :return: bool
    """
    return int(num) == float(num)


def precision_num(num: float) -> int:
    """
    Finds number of signs after comma in a float number, in case of error returns -1, in case of integer -2
    :param num: float
    :return: int
    """
    try:
        s = str(num)
        if '.' in s:
            return len(s) - s.find('.') - 1
        else:
            print('number is not float')
            return -2
    except(ValueError, TypeError):
        print('error parsing', -1)
        return -1


def is_formula(st: str) -> bool:
    """
    checks if given parameter is formula from exel
    :param st: str
    :return: bool
    """
    return True if st.startswith('=') else False


def get_end_row(col: tuple, start: int) -> int:
    """
    finds ending row in given column and sums it with start - 1 (to get last row)
    :param col: tuple
    :param start: int
    :return: int
    """
    counter = 0
    while col[counter] is not None and col[counter] != '' and counter < len(col) - 1:
        counter += 1
    return start + counter - 1


def quote_string(st: str) -> str:
    """
    add single quotes to string from both sides
    :param st: str
    :return: str
    """
    return "'{}'".format(st)


def parse_input(st: str) -> tuple:
    """
    parse user input string with path and boolean variable that says if validator should be used
    :param st:
    :return: tuple
    """
    try:
        path, set_validator = st.split(',')
        is_set = False
        if set_validator.strip() == 'y':
            is_set = True
        elif set_validator.strip() == 'n':
            is_set = False
        else:
            raise InputError
        return path, is_set
    except ValueError:
        raise InputError


# Main Class
class XLSXParser:
    STARTING_ROW = 5  # the number of row after headers
    PAGE_WITH_PERIOD_DATA = 1  # constant that stores number of page with period data
    PAGE_WITH_UNIT_DATA = 2  # constant that stores number of page with unit data
    NUM_SUBSECTIONS = ('Раз', 'Объем', 'Расценка', 'Годовая стоимость')  # needed sections for fix_num_column
    PERIOD_SUBSECTIONS = ('Периодичность', )  # needed sections for fix_other_column (applies to next line too)
    UNIT_SUBSECTIONS = ('Ед.изм.', )
    PERIOD_TRANSFER = {'раз в день, кроме праздничных и воскресных дней': 300,
                       'раз в день, кроме праздничных и выходных дней': 248,
                       'раз в квартал': 4,
                       'раз в месяц': 12,
                       'раз в неделю': 52,
                       'Круглосуточно': 365,
                       'раз в день': 365,
                       'Осмотр раз в год. По итогам осмотра работы включаются в план текущего ремонта': 1,
                       'раз в год': 1
                       }
    ENDING_ROW = 600
    ending_row = 600  # number of row to watch in first column, will be changed in __init__ according to number of rows
    filepath = 'default_name.xlsx'  # in column
    _wb = None  # variable to store .xlsx Workbook
    _ws = None  # variable to store worksheet
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')  # default error color (RED)
    yellow_fill = PatternFill(start_color='FFFFF200', end_color='FFFFF200', fill_type='solid') # default color for
    # marking period errors
    sepia_fill = PatternFill(start_color='FFE3B778', end_color='FFE3B778', fill_type='solid')
    period_list = set()  # here will be stored period data from exel file
    unit_list = set()  # here will be stored unit data from exel file
    period_validator = None  # just initializing empty variables for better readability
    unit_validator = None

    def __init__(self, path: str, is_validator: bool):
        try:
            self.filepath = path  # writes path to save
            self.is_validator = is_validator
            self._wb = load_workbook(path)  # loads .xlsx file
            self._wb.active = 0  # sets first sheet as active
            self._ws = self._wb.active  # a reference to a worksheet
            col = self._ws.iter_cols(min_col=3,
                                     max_col=3,
                                     min_row=self.STARTING_ROW,
                                     max_row=self.ending_row,
                                     values_only=True)
            self.ending_row = get_end_row(next(col), self.STARTING_ROW)
            self.period_list = self.get_values(self.PAGE_WITH_PERIOD_DATA)
            self.unit_list = self.get_values(self.PAGE_WITH_UNIT_DATA)
            # print(self.ending_row)
            if self.is_validator:
                self.period_validator = self.get_validator(self.PAGE_WITH_PERIOD_DATA)
                self.unit_validator = self.get_validator(self.PAGE_WITH_UNIT_DATA)
                self.period_validator.errorTitle = self.unit_validator.errorTitle = 'Недопустимое значение'
                self.period_validator.errorTitle = self.unit_validator.errorTitle =\
                    'Данное значение отсутствует в списке'
                self.period_validator.promptTitle = 'Выбор видов периодичности'
                self.period_validator.prompt = 'Пожалуйста выберите вид периодичности из списка'
                self.unit_validator.promptTitle = 'Выбор единиц измерения'
                self.unit_validator.prompt = 'Пожалуйста выберите единицы измерения из списка'
                self._ws.add_data_validation(self.period_validator)
                self._ws.add_data_validation(self.unit_validator)

        except (InvalidFileException, FileNotFoundError):
            print("Error! Bad path!")

    def fix_num_column(self, col: tuple, col_num: int) -> bool:
        """
        Fixes and marks errors by checking for number cell of a particular column, workbook variable must be defined,
        returns status is_modified
        :param col: tuple
        :param col_num: int
        :return: bool
        """
        is_modified = False
        if self._ws is not None:  # check is worksheet exists
            row_counter = self.STARTING_ROW  # initializing counter on the first row after header
            for cel in col:
                if cel is not None:  # checking cell conditions
                    is_num, is_changed, num = get_number(cel)
                    if is_changed:
                        is_modified = True  # if any cell should be changed return function modified status
                    if is_num:  # true
                        # print(str(cel) + ' is a number')
                        if is_integer(num):
                            # print(str(cel) + ' is integer')
                            self._ws.cell(row=row_counter, column=col_num, value=num)  # if it is int - do nothing
                        else:  # but we still assign the value to fix possible error when number has wrong separator and
                            # print(str(cel) + ' is float')
                            if precision_num(num) > 5:  # get_number() function fixes this problem
                                self._ws.cell(row=row_counter, column=col_num, value=trunc(num, 5))  # if it is float
                            else:                                   # and contains many signs after comma trunc value
                                self._ws.cell(row=row_counter, column=col_num, value=num)  # else do not touch
                    else:
                        if not is_formula(cel):
                            if self._ws.cell(row=row_counter, column=col_num).fill != self.red_fill:
                                print(str(get_column_letter(col_num)) + str(row_counter) +
                                      ' ячейка содержит ошибку с числом. Помечено красным')
                                self._ws.cell(row=row_counter, column=col_num).fill = self.red_fill  # and if it is not
                                is_modified = True
                    row_counter += 1  # a number at all mark it with marking color.
        else:
            raise NonePointer('worksheet is not defined')
        return is_modified

    def fix_other_column(self, col: tuple, col_num: int, header: str) -> bool:
        """
        Fixes period and unit column, header of the column should be passed to function to choose needed checklist,
        returns status is_modified. Can raise UndefinedHeaderError if header is wrong
        :param col: tuple
        :param col_num: int
        :param header: str
        :return: bool
        """
        is_modified = False
        if self._ws is not None:
            checklist = set()
            highlight = None
            validator = None
            if header in self.PERIOD_SUBSECTIONS:
                checklist = self.period_list
                highlight = self.yellow_fill
                validator = self.period_validator
            elif header in self.UNIT_SUBSECTIONS:
                checklist = self.unit_list
                highlight = self.sepia_fill
                validator = self.unit_validator
            else:
                raise UndefinedHeaderError('header does not exist in any of given subsections')
            row_counter = self.STARTING_ROW
            for cel in col:
                if cel not in checklist and cel is not None and cel != '':
                    if self._ws.cell(row_counter, col_num).fill != highlight:  # if not yet marked
                        if checklist is self.unit_list:
                            print(str(get_column_letter(col_num)) + str(row_counter) +
                                  ' ячейка содержит ошибку единицы измерения. Помечено цветом сепии')
                        else:
                            print(str(get_column_letter(col_num)) + str(row_counter) +
                                  ' ячейка содержит ошибку вида периодичности. Помечено желтым')
                        self._ws.cell(row_counter, col_num).fill = highlight
                        is_modified = True  # if any cell highlighted (changed) - change modify status
                row_counter += 1
            if self.is_validator:
                validator.add(str(get_column_letter(col_num)) + str(self.STARTING_ROW) + ':'
                              + str(get_column_letter(col_num)) + str(self.ENDING_ROW))
        else:
            raise NonePointer('worksheet is not defined')
        return is_modified

    def find_errors(self):
        """
        this method parses the table and fixes the errors with fix_num_column and fix_other_column. Also
        it checks value in 'Годовая стоимость' and replaces it if it contains error
        :return:
        """
        if self._ws is not None:
            is_num_modified = False
            is_other_modified = False
            col_counter = 4
            subsection_amount = 961
            n = []
            periodicity = []
            amount = []
            tariff = []
            print('Если таблица не содержит ошибок или они уже были помечены - ничего не будет выведено в лог'
                  ', иначе будут выведены номера ячеек с ошибками, типом ошибки и помеченным цветом')
            for i in range(subsection_amount):
                header = self._ws.cell(column=col_counter, row=self.STARTING_ROW - 1).value
                col = self._ws.iter_cols(min_col=col_counter,
                                         max_col=col_counter,
                                         min_row=self.STARTING_ROW,
                                         max_row=self.ending_row,
                                         values_only=True)
                col_tup = next(col)
                # print(self._ws.cell(column=col_counter, row=self.STARTING_ROW - 1).value)
                if header in self.NUM_SUBSECTIONS:
                    tmp = self.fix_num_column(col_tup, col_counter)
                    if tmp:
                        is_num_modified = True
                    # print(col_tup)
                elif header in self.UNIT_SUBSECTIONS or header in self.PERIOD_SUBSECTIONS:
                    tmp = self.fix_other_column(col_tup, col_counter, header)
                    if tmp:
                        is_other_modified = True
                print(header)
                if header == self.NUM_SUBSECTIONS[0]:
                    n = get_column_letter(col_counter)
                    print(n)
                if header == self.PERIOD_SUBSECTIONS[0]:
                    periodicity = self.trans_period(col_tup)
                    print(periodicity)
                if header == self.NUM_SUBSECTIONS[1]:
                    amount = get_column_letter(col_counter)
                    print(amount)
                if header == self.NUM_SUBSECTIONS[2]:
                    tariff = get_column_letter(col_counter)
                    print(tariff)
                if header == self.NUM_SUBSECTIONS[3] and len([i for i in n if i is not None]) > 0 and \
                        len([i for i in periodicity if i is not None]) > 0 and \
                        len([i for i in amount if i is not None]) > 0 and \
                        len([i for i in tariff if i is not None]) > 0:
                    price = self.form_price(periodicity, n, amount, tariff)
                    self.assign_col(price, col_counter)
                col_counter += 1
            # print(is_num_modified, is_other_modified)
            if is_num_modified or is_other_modified or self.is_validator:
                self._wb.save(self.filepath)
        else:
            raise NonePointer('Worksheet is not defined')

    def get_values(self, sheet: int) -> tuple:
        """
        Scans values from the first column of given Worksheet and returns them in tuple
        :param sheet: int
        :return: tuple
        """
        if self._wb is not None:
            try:
                self._wb.active = sheet
                ws = self._wb.active
                counter = 1
                ans = []
                while ws.cell(counter, 1).value is not None and ws.cell(counter, 1).value != '' and counter <= 1000:
                    ans.append(ws.cell(counter, 1).value)
                    counter += 1
                self._wb.active = 0  # sets first page as active after all actions, just in case
                ans = tuple(OrderedDict.fromkeys(ans))  # delete all duplicates preserving order
                # print(ans)
                return ans
            except AttributeError:
                print('page is not defined or does not exist')
        else:
            raise NonePointer('Workbook is not defined')

    def get_validator(self, page: int) -> DataValidation:
        """
        returns data validator initialized with specified list(set)
        :param page: int (number of page with data that needs to be validated)
        :return:
        """
        lt = self.get_values(page)  # this is needed just to find number of rows in the next line
        end_row = get_end_row(lt, 1)
        self._wb.active = page
        # print("{}!$A$1:$A${}".format(quote_string(self._wb.active.title), end_row))
        dv = DataValidation(type='list', formula1="{}!$A$1:$A${}".format(quote_string(self._wb.active.title), end_row),
                            showDropDown=False)
        self._wb.active = 0  # return to zero page
        return dv

    def trans_period(self, col: tuple) -> tuple:
        """
        transfers string in column 'Периодичность' to a number according to PERIOD_TRANSFER constant
        :param col: tuple (a tuple of a column "Периодичность")
        :return: tuple (with numbers according to PERIOD_TRANSFER
        """
        ans = []
        for el in col:
            ans.append(self.PERIOD_TRANSFER.get(el))
        return tuple(ans)

    def assign_col(self, col: tuple, col_num: int):
        """
        assign values from the tuple to a specified column
        :param col: tuple (tuple with data to assign)
        :param col_num: int (number of column to assign)
        :return:
        """
        row_counter = self.STARTING_ROW
        for el in col:
            if el != '#ERR':
                self._ws.cell(row_counter, col_num, el)
            row_counter += 1

    def form_price(self, periodicity: tuple, n: str, amount: str, tariff: str) -> tuple:
        """returns formula tuple of 'Годовая стоимость' column"""
        ans = []
        row_counter = self.STARTING_ROW
        for period in periodicity:
            if period is not None and self._ws.cell(row_counter, column_index_from_string(n)).value is not None and \
                    self._ws.cell(row_counter, column_index_from_string(amount)).value is not None and \
                    self._ws.cell(row_counter, column_index_from_string(tariff)).value is not None:
                ans.append('=' + n + str(row_counter) + '*' + str(period) + '*' + amount + str(row_counter) +
                           '*' + tariff + str(row_counter) + '/1000')
            else:
                ans.append('#ERR')
            row_counter += 1
        return tuple(ans)


s = input('Введите путь к .xlsx файлу и через запятую напишите y/n нужно/не нужно добавлять валидатор: ')
# print(s)
path, is_validator = parse_input(s)
xl = XLSXParser(path, is_validator)
xl.find_errors()
