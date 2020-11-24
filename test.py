from collections import OrderedDict

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

wb = load_workbook('example_form_2_3.xlsx')
wb.active = 0
ws = wb.active
ws['R7'] = 82
print(ws.cell(7, column_index_from_string('R')).value)
wb.save('example_form_2_3.xlsx')